import openpyxl
import re
import json

def create_flatten_sheet(wb):
	for name in wb.get_sheet_names():
		if name == "data_flatten":
			wb.remove_sheet(wb.get_sheet_by_name('data_flatten'))
	wb.create_sheet('data_flatten')

def isDateRow(date_str):
	return re.compile('Week').match(date_str) != None

def get_date_values(date_str):
	q_string = 'Week (\d+).+(201[678]).+(\d+)-(\d+)'
	date_re = re.compile(q_string).match(date_str)
	curr_week = date_re.group(1)
	curr_date = date_re.group(3) + "/" + date_re.group(4) + "/" + date_re.group(2)
	return curr_week, curr_date

'''
def isProductRow(col_1_val, re_string):
	#return re.match( r'(.*)oz', col_1_val) != None
	return re.compile(re_string).match(col_1_val) != None
'''

def isProductRow(col_1_val):
	return re.compile('^\s{2}\S+').match(col_1_val) != None


#
# re_string is deprecated in processSheet() but I don't want to remove it
# until we're sure it is not needed
#
def processSheet(workbook, sheet_name, re_string, output_data):
	sheet = workbook.get_sheet_by_name(sheet_name)
	max_rows = sheet.max_row

	curr_product = "Unknown"
	curr_week = "Unknown"
	curr_date = "Unknown"
	for row in range(1, sheet.max_row+1):
		col_1_val = sheet['A' + str(row)].value
		if isDateRow(col_1_val):
			curr_week, curr_date = get_date_values(col_1_val)
		elif isProductRow(col_1_val):
			curr_product = col_1_val
		else:
			sku = col_1_val
			num_orders = str(sheet['B' + str(row)].value)
			amount = str(sheet['C' + str(row)].value)
			entry = { "name":sheet_name, "product":curr_product, "sku":sku, "num_orders":num_orders, "amount":amount, "week":curr_week, "date":curr_date}
			output_data.append(entry)
			print(curr_product + "," + sku + "," + num_orders + "," + amount)

def print_headings(sheet):
	sheet['A1'] = "category"
	sheet['B1'] = "product"
	sheet['C1'] = "sku"
	sheet['D1'] = "num_orders"
	sheet['E1'] = "amount"
	sheet['F1'] = "week"
	sheet['G1'] = "date"

wb = openpyxl.load_workbook('data.xlsx')
output_data = []
processSheet(wb, "mug", '(.*)oz', output_data)
processSheet(wb, "album", '(.*)[Photo|Layflat|Cover] Book|(.*)Flipbook', output_data)
processSheet(wb, "calendar", '(.*)Calendar', output_data)
processSheet(wb, "blanket", '(.*)Blanket|(.*)Throw|(.*)Quilt|(.*)Tied\ Up', output_data)
processSheet(wb, "blanket", '(.*)Blanket|(.*)Throw|(.*)Quilt|(.*)Tied\ Up', output_data)
processSheet(wb, "card", 'placeholder', output_data)


wb = openpyxl.load_workbook('data.xlsx')
create_flatten_sheet(wb)
sheet = wb.get_sheet_by_name('data_flatten')
print_headings(sheet)
#print str(sheet)

i=2

for entry in output_data:
	#print( str(entry))
	sheet['A' + str(i)] = entry["name"]
	sheet['B' + str(i)] = entry["product"]
	sheet['C' + str(i)] = entry["sku"]
	sheet['D' + str(i)] = int(entry["num_orders"])
	sheet['E' + str(i)] = float(entry["amount"])
	sheet['F' + str(i)] = int(entry["week"])
	sheet['G' + str(i)] = entry["date"]
	i+= 1
wb.save("data.xlsx")

output_data_json = json.dumps(output_data)
flatten_json_file = open("data_flatten.json", 'w')
flatten_json_file.write(output_data_json)
flatten_json_file.close()








